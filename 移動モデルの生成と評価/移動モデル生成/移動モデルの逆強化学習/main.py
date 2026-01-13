import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import numpy as np
from multiprocessing import Pool
import time
from tqdm import tqdm
from value_interation import ValueIteration
import math
import openpyxl as excel
experiment_name="ex1"
class maxent():
    def __init__(self,hour):
        self.n_states = None
        self.n_actions = None
        self.value_iteration=None
        self.expart_features1=None
        self.hour=hour
    def irl(self,trajectories,states,n):
        
        self.n_states=1600
        #print(1)

        #状態数の行列の中身をランダムで埋める(0~1)初期報酬となる
        theta = np.random.uniform(size=states)
        #print(2)
        
        #エキスパートの軌道ζを特徴ベクトルFζに変換
        expart_features = self.onehot_feature(trajectories,states)#エキスパート軌道から特徴ベクトルを計算1/M*Σfs
        #print(3)
        self.expart_features1=expart_features
        #print(4)
        #状態ごとの単位ベクトル
        state_features = np.vstack([self.state_to_feature2(s)
                                   for s in range(self.n_states)])
        #print(5)
                                   

        ac=int(math.sqrt(len(expart_features)))
        #print(6)
        st=np.zeros((ac,ac),dtype=np.float64)
        #print(7)
        for i in range(ac):
            for j in range(ac):
                #print(teacher_features[i*ac+j])
                #print(8)
                st[i][j]=expart_features[i*ac+j]
        print("plot_Expart")
        #print(9)
        plt.close()

        plt.figure()
        #print(10)
        fig, ax = plt.subplots()
        #print(11)
        plt.pcolor(st[::-1, :])
        #print(12)
        plt.colorbar()
        #print(13)
        
        plt.title("エキスパートの特徴量", fontname="MS Gothic")
        #print(14)
        plt.savefig('./'+str('./expart_feature.png'))
        #print(15)
        plt.close()
        #print(16)
        
        self.value_iteration = ValueIteration(states, 4)#16 4,list
        #print(17)
        self.n_actions=4
        #print(18)
        self.n_states=states
        #print(19)
        learning_rate=0.1
        #print(20)
        
        for e in tqdm(range(n)):
            #各状態の報酬をθと平均訪問回数のドット積で計算>R(ζ)=θT*fζ
            #reward=state_features*theta
            #print(1)
            reward=state_features.dot(theta.T)
            #現在の報酬からえられる方策を計算
            #print(2)
            self.value_iteration.set_reward(reward)
            #print(3)
            policy=self.value_iteration.reward2policy()
            #print(4)
            #推定した方策の特徴ベクトルを計算
            #print(55555)
            #print(len(trajectories))
            features = self.policy2feature(policy, trajectories)
            #print(5)
            #勾配を計算
            theta+=learning_rate * (expart_features - features.dot(state_features))
            #print(6)
            #途中経過を記録
            estimated = state_features.dot(theta.T)
            #print(7)
            estimated = estimated.reshape((40,40))
            #print(8)
            plt.close()
            #print(9)
        
        
            plt.figure()
            fig, ax = plt.subplots()
            plt.pcolor(estimated[::-1, :])
            plt.colorbar()
            
            plt.title("IRL"+str(e)+"回目で予測される報酬", fontname="MS Gothic")
            plt.savefig('./'+str('./reward'+str(e))+'.png')
            plt.close()
            #self.grid_plot(array= estimated,title="IRL"+str(e)+"回目で予測される報酬",name='./reward'+str(e))
            #print(2)
            
        #報酬を計算
        estimated = state_features.dot(theta.T)#
        estimated = estimated.reshape((40,40))
        plt.close()
        #print(3)
    
    
        plt.figure()
        fig, ax = plt.subplots()
        plt.pcolor(estimated[::-1, :])
        plt.colorbar()
        #print(4)
        
        plt.title("IRLで予測される報酬", fontname="MS Gothic")
        plt.savefig('./'+str('./rewardresult.png'))
        self.list2excel(estimated,expart_features,"result"+str(self.hour))
        plt.close()
        #print(5)
    
    def state_to_feature(self, s):
        feature = np.zeros(self.n_states)
        feature[s] = 1
        #print(6)
        return feature
    
    def state_to_feature2(self, s):
        feature = np.zeros(self.n_states)
        feature[s] = self.expart_features1[s]
        #print(7)
        return feature
    
    def grid_plot(array,title,name,excelsave=False):
        plt.close()
        
        
        plt.figure()
        fig, ax = plt.subplots()
        plt.pcolor(array[::-1, :])
        plt.colorbar()
        
        plt.title(title, fontname="MS Gothic")
        plt.savefig('./'+str(name)+'.png')
        plt.close()
        #print(8)

        # if excelsave==True:
        #     list2excel(array,name)

    def list2excel(self,array,array2,name):
        book = excel.Workbook()
        sheet = book.active
        #print(type(array))
        for i in range(len(array)):
            for j in range(len(array)):
                sheet.cell(row=(i*len(array))+j+1, column=1, value=(i*len(array))+j+1)#grid No
                sheet.cell(row=(i*len(array))+j+1, column=2, value=float(array[i,j]))
                sheet.cell(row=(i*len(array))+j+1, column=3, value=float(array2[i*40+j]))
        book.save(str(name)+".xlsx")
        #print(9)

    def onehot_feature(self,trajectories,states):
        features = np.zeros(states)#特徴量ベクトルを宣言(use env)
        #print(1)
        for i in trajectories:#各軌道データごとで回す
            #print(2)
            for s in i:#各軌道から1stepごと
                #print(3)
                features[s] += 1#0~nまでの訪問回数をカウント

        #print(len(trajectories))
        #//////////
        features /= len(trajectories)#平均訪問回数を計算(特殊な計算式だが、リスト内の各要素に対して割り算を行っているだけ)1/M*Σfs
        #/////////
        #print(4)
        return features#平均訪問回数を返えす

    def policy2feature(self,policy,trajectories):
        # エキスパート軌跡の数, このコードではt_size=40
        #print(1)
        #print(len(trajectories))
        t_size = len(trajectories)#エキスパートの軌道の数
        #print(2)
        states = self.n_states #各状態を呼び出し(list)
        #print(3)

        # 状態遷移確率 軌跡の数×状態数のnplistを作成
        transition_probs = np.zeros((t_size, states))
        #print(4)

        # 状態の発生確率(各状態の到達頻度)
        initial_state_probs = np.zeros(states)
        #print(5)
        # 各軌跡データの初期状態を取得して、数をカウント
        for t in trajectories:
            initial_state_probs[t[0]] += 1
            #print(6)
        # 回数を頻度に変換するためにt_sizeで割る
        #print(len(t_size))
        #////////
        initial_state_probs /= t_size
        #/////////
        #print(7)
        # 状態遷移確率の初期行に初期状態の配列を代入
        transition_probs[0] = initial_state_probs
        #print(8)

        # 環境の状態遷移確率にしたがって,t_size-1回状態遷移を繰り返して状態の発生確率を計算
        for t in range(1, t_size):
            #print(9)
            # 1ステップ前の状態の発生確率と状態遷移確率を掛けて全て足すことで今ステップの状態発生確率を計算
            # μ_t(s') = P(s'|s, a) * Σμ_t-1(s)
            #transition_probs[t][s] = transition_probs[t - 1][prev_s] * self.env.transit_func(prev_s, self.planner.act(prev_s))
            for prev_s in range(states):#各状態から次のステップをまわす
                # 1ステップ前に状態prev_sにいる確率を計算
                #print(1)
                prev_prob = transition_probs[t - 1][prev_s]
                # 1ステップ前の状態prev_sで行う行動を方策から決定
                #print(2)
                a = np.argmax(prev_s)#self.planner.act(prev_s)#int型意味➡方策の中で最も報酬の高い移動先を算出
                # 状態遷移確率に従い、各状態への遷移確率を取得. probsは16次元のリスト
                probs = self.value_iteration.transit_func(prev_s, a)
                # 1ステップ前の発生確率と状態遷移確率を掛ける
                #print(3)
                for s in probs:#各状態を保存
                    transition_probs[t][int(s)] += prev_prob * probs[int(s)]
                    #print(4)

        # t_sizeステップ分の発生確率を平均して、各状態の発生確率を計算
        total = np.mean(transition_probs, axis=0)
        return total
        
    
#def multicalc(x,y):
def multicalc(inputs):
    import db2expart
    import os
    import sys
    import random
    #print(55555)
    #print(inputs)
    hour=inputs
    #print(55555)
    #print(hour)
    #print(1)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    #print(2)
    try:

        trajectories,states=db2expart.db2trajectory(hour)
    except:   
        return
    #print(3)
    #print(55555)
    #print(len(trajectories))
    #print(4)
    #print (sys.argv[1],sys.argv[2])
    os.makedirs('result/'+str(experiment_name)+'/'+str(hour), exist_ok=True)#実験名/時刻でフォルダを作成
    #print(5)
    os.makedirs('result/'+str(experiment_name)+'/result', exist_ok=True)#実験名/時刻でフォルダを作成
    #print(6)
    os.chdir('result/'+str(experiment_name)+'/'+str(hour))
    #print(7)



    print("start IRL")
    Irl=maxent(hour=hour)
    #print(7)

    Irl.irl(trajectories,states, 200)
    #print(8)
    return

         
if __name__ == "__main__":
    import db2expart
    import os
    import sys
    import random
    p = Pool(4)
    #print(1)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    #print(2)
    values = [x for x in range(0,24) ]
    #print(3)
    
    p.map(multicalc, [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24 ])
    #print(4)
    # print(result)

            
# if __name__ == "__main__":
#     import db2expart
#     import os
#     import sys
#     import random
#     os.chdir(os.path.dirname(os.path.abspath(__file__)))
    
#     #実験名とIRLに入力する時刻を決定
#     try:
#         sys.argv[1]
#         experiment_name=sys.argv[1]
#     except:
#         experiment_name=random.randrange(10000)
#     try:
#         hour=sys.argv[2]
#     except:
#         hour=12
    
#     trajectories,states=db2expart.db2trajectory(hour)
#     os.chdir(os.path.dirname(os.path.abspath(__file__)))
#     #print (sys.argv[1],sys.argv[2])
#     os.makedirs('result/'+str(experiment_name)+'/'+str(hour), exist_ok=True)#実験名/時刻でフォルダを作成
#     os.makedirs('result/'+str(experiment_name)+'/result', exist_ok=True)#実験名/時刻でフォルダを作成
#     os.chdir('result/'+str(experiment_name)+'/'+str(hour))



#     print("start IRL")
#     Irl=maxent()

#     Irl.irl(trajectories,states, 150)

    
